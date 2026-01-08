#!/usr/bin/env python3
"""
Script to generate Excel file from TEST_CASES.md
"""

import csv
import re
from datetime import datetime

def parse_test_cases():
    """Parse test cases from TEST_CASES.md"""
    test_cases = []
    
    with open('TEST_CASES.md', 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Find all test case sections
    sections = re.findall(r'## Module: (.+?)\n\n\|.*?\n\|.*?\n(.*?)(?=\n---|\n## |$)', content, re.DOTALL)
    
    for module_name, table_content in sections:
        # Extract table rows
        rows = re.findall(r'\| (WA\d+) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \| (.+?) \|', table_content)
        
        for row in rows:
            test_case = {
                'Test Case ID': row[0],
                'Module': row[1],
                'Feature / Screen': row[2],
                'Test Scenario': row[3],
                'Test Steps': row[4].replace('<br>', '\n'),
                'Test Data': row[5].replace('<br>', '\n'),
                'Expected Result': row[6].replace('<br>', '\n'),
                'Actual Result': row[7],
                'Status (Pass/Fail/Blocked)': row[8],
                'Severity (Low/Medium/High/Critical)': row[9],
                'Defect ID': row[10],
                'Remarks': row[11]
            }
            test_cases.append(test_case)
    
    return test_cases

def write_to_csv(test_cases, filename='Test_Cases.csv'):
    """Write test cases to CSV file"""
    if not test_cases:
        print("No test cases found!")
        return
    
    fieldnames = [
        'Test Case ID',
        'Module',
        'Feature / Screen',
        'Test Scenario',
        'Test Steps',
        'Test Data',
        'Expected Result',
        'Actual Result',
        'Status (Pass/Fail/Blocked)',
        'Severity (Low/Medium/High/Critical)',
        'Defect ID',
        'Remarks'
    ]
    
    with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(test_cases)
    
    print(f"✓ Successfully created {filename} with {len(test_cases)} test cases")

def write_to_excel(test_cases, filename='Test_Cases.xlsx'):
    """Write test cases to Excel file using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Headers
        headers = [
            'Test Case ID',
            'Module',
            'Feature / Screen',
            'Test Scenario',
            'Test Steps',
            'Test Data',
            'Expected Result',
            'Actual Result',
            'Status (Pass/Fail/Blocked)',
            'Severity (Low/Medium/High/Critical)',
            'Defect ID',
            'Remarks'
        ]
        
        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Write data
        for row_num, test_case in enumerate(test_cases, 2):
            for col_num, header in enumerate(headers, 1):
                value = test_case.get(header, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border
                
                # Color code by severity
                if header == 'Severity (Low/Medium/High/Critical)':
                    if value == 'Critical':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif value == 'High':
                        cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif value == 'Medium':
                        cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
                    elif value == 'Low':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                
                # Color code by status
                if header == 'Status (Pass/Fail/Blocked)':
                    if value == 'Pass':
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif value == 'Fail':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif value == 'Blocked':
                        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
        
        # Adjust column widths
        column_widths = {
            'A': 12,  # Test Case ID
            'B': 20,  # Module
            'C': 25,  # Feature / Screen
            'D': 35,  # Test Scenario
            'E': 50,  # Test Steps
            'F': 30,  # Test Data
            'G': 40,  # Expected Result
            'H': 40,  # Actual Result
            'I': 25,  # Status
            'J': 30,  # Severity
            'K': 15,  # Defect ID
            'L': 30   # Remarks
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        wb.save(filename)
        print(f"✓ Successfully created {filename} with {len(test_cases)} test cases")
        
    except ImportError:
        print("⚠ openpyxl not installed. Creating CSV instead...")
        write_to_csv(test_cases, filename.replace('.xlsx', '.csv'))

if __name__ == '__main__':
    print("Generating test cases Excel file...")
    test_cases = parse_test_cases()
    
    if test_cases:
        # Try to create Excel, fallback to CSV
        try:
            write_to_excel(test_cases)
        except:
            write_to_csv(test_cases)
    else:
        print("No test cases found in TEST_CASES.md")

